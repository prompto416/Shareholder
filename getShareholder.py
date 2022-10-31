from openpyxl import load_workbook

wb = load_workbook(filename='SET symbol and sector.xlsx',read_only=True)

ws = wb['SET']

for row in ws['A2:Z999']:
    for cell in row:
        if (cell.value != None) and (cell.value != 'Sector') and( cell.value != 'Symbol'):
            print(cell.value)
            
        